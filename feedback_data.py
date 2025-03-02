"""
feedback_data.py

This module handles the processing of feedback data from Excel files.
It provides functions to load, filter, and transform feedback data.

Dependencies:
    - pandas
    - openpyxl
    - numpy
"""

# Third-party libraries
import pandas as pd
import openpyxl as xl
import numpy as np

# Built-in libraries
import re
import os
import glob
import traceback
from typing import Tuple, List, Optional, Union

def get_feedback_files(main_directory: str) -> List[str]:
    """
    Find all feedback Excel files in the specified directory and its subdirectories.
    
    Args:
        main_directory (str): Root directory to search for feedback files
        
    Returns:
        List[str]: List of paths to feedback files
    """
    pattern = os.path.join(main_directory, '**', '*.xlsm')
    xlsm_files = glob.glob(pattern, recursive=True)
    return [file for file in xlsm_files if os.path.basename(file).startswith('feedback')]

def extract_sme_code(filename: str) -> Optional[str]:
    """
    Extract SME code from filename.
    
    Args:
        filename (str): Name of the file to extract SME code from
        
    Returns:
        Optional[str]: Extracted SME code or None if not found
    """
    pattern = r'EVAL-(?:consensus|\d+)'
    match = re.search(pattern, filename)
    return match.group() if match else None

def dimension_index(x: Union[str, float]) -> str:
    """
    Process dimension values to standardized format.
    
    Args:
        x: Input dimension value
        
    Returns:
        str: Processed dimension value
    """
    if pd.isna(x) or x == '' or x == 'n/a':
        return 'na'
    return x.split('--')[0].strip()

def load_raw_feedback(datapathlist: List[str]) -> Tuple[Optional[pd.DataFrame], Optional[pd.DataFrame]]:
    """
    Load and process feedback data from Excel files.
    
    Args:
        datapathlist (List[str]): List of paths to feedback files
        
    Returns:
        Tuple[Optional[pd.DataFrame], Optional[pd.DataFrame]]: 
            Tuple of (QA data, reference data) DataFrames
    """
    try:
        qa_data_list = []
        ref_data_list = []
        
        for path in datapathlist:
            sme_name = extract_sme_code(path)
            book = xl.load_workbook(path)
            
            if 'Feedback' in book.sheetnames:
                df_qa = pd.read_excel(path, sheet_name='Feedback')
                headers = df_qa.iloc[0]
                df_qa = pd.DataFrame(df_qa.values[1:], columns=headers)
                df_qa['SME'] = sme_name
                df_qa = df_qa.dropna(subset=['Query ID'])
                qa_data_list.append(df_qa)
                
            if 'References' in book.sheetnames:
                df_ref = pd.read_excel(path, sheet_name='References')
                df_ref['SME'] = sme_name
                df_ref = df_ref.dropna(subset=['Query ID'])
                ref_data_list.append(df_ref)
                
            book.close()  # Properly close workbook
            
        combined_qa = pd.concat(qa_data_list, ignore_index=True) if qa_data_list else pd.DataFrame()
        combined_ref = pd.concat(ref_data_list, ignore_index=True) if ref_data_list else pd.DataFrame()
        
        return combined_qa, combined_ref
        
    except Exception as e:
        print(f"Error in load_raw_feedback: {str(e)}")
        traceback.print_exc()
        return None, None

def get_reviewed_queries(feedback: pd.DataFrame) -> Optional[pd.DataFrame]:
    """
    Filter for reviewed queries meeting specific criteria.
    
    Args:
        feedback (pd.DataFrame): Input feedback data
        
    Returns:
        Optional[pd.DataFrame]: Filtered feedback data
    """
    try:
        # Filter out unreviewed queries
        feedback = feedback[feedback['Unable to Review'] != 'X']
        
        # Remove entries missing critical ratings
        feedback = feedback.dropna(subset=[
            'Overall Answer Helpfulness',
            'Comprehension',
            'Clinical Harmfulness'
        ])
        
        # Filter comprehension scores
        comprehend = feedback[feedback['Comprehension'] != '0']
        
        # Handle 'na' values
        na_queries = set(
            list(comprehend[comprehend['Correctness'] == 'na']['Query ID']) +
            list(comprehend[comprehend['Completeness'] == 'na']['Query ID'])
        )
        
        return feedback[~feedback['Query ID'].isin(na_queries)]
        
    except Exception as e:
        print(f"Error in get_reviewed_queries: {str(e)}")
        traceback.print_exc()
        return None

def get_unable_to_review_queries(feedback: pd.DataFrame) -> Optional[pd.DataFrame]:
    """
    Filter for queries marked as unable to review.
    
    Args:
        feedback (pd.DataFrame): Input feedback data
        
    Returns:
        Optional[pd.DataFrame]: Filtered feedback data
    """
    try:
        return feedback[feedback['Unable to Review'] == 'X']
    except Exception as e:
        print(f"Error in get_unable_to_review_queries: {str(e)}")
        traceback.print_exc()
        return None

def convert_to_dimensionscore(feedback: pd.DataFrame) -> Optional[pd.DataFrame]:
    """
    Convert feedback ratings to numerical dimension scores.
    
    Args:
        feedback (pd.DataFrame): Input feedback data
        
    Returns:
        Optional[pd.DataFrame]: Processed feedback data with numerical scores
    """
    try:
        # Create a copy to avoid modifying original DataFrame
        feedback = feedback.copy()
        
        # Convert emoji ratings to numerical scores
        emoji_map = {
            ' 🙁': 0,
            ' 😐': 1,
            ' 😀': 2
        }
        
        feedback['Overall Answer Helpfulness'] = feedback['Overall Answer Helpfulness'].map(
            emoji_map).fillna(feedback['Overall Answer Helpfulness'])
        
        # Process dimension columns
        dimension_columns = [
            'Comprehension',
            'Correctness',
            'Completeness',
            'Clinical Harmfulness',
            'Clinical Harmfulness Level'
        ]
        
        for col in dimension_columns:
            feedback[col] = feedback[col].apply(dimension_index)
            
        return feedback
        
    except Exception as e:
        print(f"Error in convert_to_dimensionscore: {str(e)}")
        traceback.print_exc()
        return None
